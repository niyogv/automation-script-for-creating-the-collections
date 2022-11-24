# This a script which reads data from excel sheet

import openpyxl



class Home:

    @staticmethod
    def readDataLogin(test_case_name): # login test data
        book=openpyxl.load_workbook('/Users/aicumen-dev/Downloads/data.xlsx')
        dict={}
        sheet=book.get_sheet_by_name('login')
        rcount=sheet.max_row
        Ccount=sheet.max_column

        for r in range(1,rcount+1):
            if sheet.cell(row=r,column=1).value==test_case_name:
               for c in range(2,Ccount+1):
                   dict[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value
        return [dict]

    @staticmethod
    def readDataCollection(test_case_name): # collection creation data
        book=openpyxl.load_workbook('/Users/aicumen-dev/Downloads/data.xlsx')
        dict={}
        sheet=book.get_sheet_by_name('collection')
        rcount=sheet.max_row
        Ccount=sheet.max_column

        for r in range(1,rcount+1):
            if sheet.cell(row=r,column=1).value==test_case_name:
                for c in range(2,Ccount+1):
                    dict[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value
        return [dict]

    @staticmethod
    def readDataList(test_case_name): # data of list your edition after minting
        book=openpyxl.load_workbook('/Users/aicumen-dev/Downloads/data.xlsx')
        dict={}
        sheet=book.get_sheet_by_name('list_ur_edition')
        rcount=sheet.max_row
        Ccount=sheet.max_column

        for r in range(1,rcount+1):
            if sheet.cell(row=r,column=1).value==test_case_name:
                for c in range(2,Ccount+1):
                    dict[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value
        return [dict]

    @staticmethod
    def readDataMint(test_case_name): # data of list your edition after minting
        book=openpyxl.load_workbook('/Users/aicumen-dev/Downloads/data.xlsx')
        dict={}
        sheet=book.get_sheet_by_name('mint')
        rcount=sheet.max_row
        Ccount=sheet.max_column

        for r in range(1,rcount+1):
            if sheet.cell(row=r,column=1).value==test_case_name:
                for c in range(2,Ccount+1):
                    dict[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value
        return [dict]

